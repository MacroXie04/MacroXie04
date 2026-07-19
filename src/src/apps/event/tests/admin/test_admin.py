import json
import uuid
from io import BytesIO
from unittest.mock import patch

from django.contrib.admin.helpers import ACTION_CHECKBOX_NAME
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from apps.authn.models import ContactEmail, ContactPhone, Member
from apps.event.admin.registration import EventRegistrationAdmin
from apps.event.models import CheckIn, CheckInRecord, Event, EventRegistration, Question, Ticket
from apps.event.services import ScheduleSyncStats, build_event_copy_template
from apps.event.tests.helpers import (
    make_event,
    make_member,
    make_registration,
    make_superuser,
    make_ticket,
)


class EventAdminTest(TestCase):
    def setUp(self):
        self.admin_user = make_superuser()
        self.client.login(username="admin@example.com", password="testpass123")

    @staticmethod
    def _copy_post_data(*, name, slug, ticket_rows=(), question_rows=(), **overrides):
        data = {
            "name": name,
            "slug": slug,
            "date": "2026-10-01",
            "end_date": "2026-10-03",
            "location": "Copied Hall",
            "description": "Reviewed copied Event.",
            "ticket_login_validity_days": "30",
            "ticket_login_reusable": "on",
            "registration_sheet_id": "",
            "registration_sheet_gid": "",
            "tickets-TOTAL_FORMS": str(len(ticket_rows)),
            "tickets-INITIAL_FORMS": "0",
            "tickets-MIN_NUM_FORMS": "0",
            "tickets-MAX_NUM_FORMS": "1000",
            "questions-TOTAL_FORMS": str(len(question_rows)),
            "questions-INITIAL_FORMS": "0",
            "questions-MIN_NUM_FORMS": "0",
            "questions-MAX_NUM_FORMS": "1000",
            "_save": "Save",
        }
        for index, row in enumerate(ticket_rows):
            data[f"tickets-{index}-name"] = row["name"]
            data[f"tickets-{index}-order"] = str(row["order"])
            if row.get("DELETE"):
                data[f"tickets-{index}-DELETE"] = "on"
        for index, row in enumerate(question_rows):
            data[f"questions-{index}-text"] = row["text"]
            data[f"questions-{index}-order"] = str(row["order"])
            if row.get("is_required"):
                data[f"questions-{index}-is_required"] = "on"
            if row.get("DELETE"):
                data[f"questions-{index}-DELETE"] = "on"
        data.update(overrides)
        return data

    def test_changelist_accessible(self):
        response = self.client.get("/admin/event/event/")
        self.assertEqual(response.status_code, 200)

    def test_add_form_accessible(self):
        response = self.client.get("/admin/event/event/add/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Copy from existing Event")
        self.assertContains(response, "event/js/event_admin.js")
        self.assertContains(response, "event/css/event_admin.css")
        self.assertContains(response, "Prompt for Phone Number")
        self.assertContains(response, 'id="event-copy-source-hint"')
        self.assertContains(response, 'aria-describedby="event-copy-source-hint"')
        self.assertContains(response, 'aria-hidden="true"')
        self.assertContains(response, 'id="event-verify-phone-dependency-hint"')
        self.assertContains(response, "event-verify-phone-dependency-hint", count=2)

    def test_copy_template_builder_returns_ordered_safe_snapshot(self):
        source = make_event(
            name="Builder Source",
            registration_open=True,
            registration_sheet_id="private-source-sheet",
            registration_sheet_gid=72,
        )
        later_ticket = Ticket.objects.create(event=source, name="VIP", order=20)
        first_ticket = Ticket.objects.create(event=source, name="General", order=10)
        later_question = Question.objects.create(event=source, text="Dietary needs?", order=9)
        first_question = Question.objects.create(event=source, text="Role?", is_required=True, order=2)

        template = build_event_copy_template(source)

        self.assertEqual(template.source_id, source.pk)
        self.assertEqual(template.source_name, source.name)
        self.assertEqual(template.event_initial["name"], "")
        self.assertEqual(template.event_initial["slug"], "")
        self.assertFalse(template.event_initial["registration_open"])
        self.assertEqual(template.event_initial["registration_sheet_id"], "")
        self.assertIsNone(template.event_initial["registration_sheet_gid"])
        self.assertEqual(
            template.ticket_initial(),
            [
                {"name": first_ticket.name, "order": 10},
                {"name": later_ticket.name, "order": 20},
            ],
        )
        self.assertEqual(
            template.question_initial(),
            [
                {"text": first_question.text, "is_required": True, "order": 2},
                {"text": later_question.text, "is_required": False, "order": 9},
            ],
        )
        self.assertNotIn("barcode", template.ticket_initial()[0])

    def test_add_form_loads_safe_copy_source_without_creating_event(self):
        source = make_event(
            name="Source Event",
            date="2026-05-14",
            end_date="2026-05-16",
            location="Source Hall",
            description="Reusable source description.",
            registration_open=True,
            allow_secondary_email=True,
            collect_phone=True,
            verify_phone=True,
            ticket_login_validity_days=45,
            ticket_login_reusable=False,
            registration_sheet_id="source-sheet",
            registration_sheet_gid=987,
            registration_sheet_sync_count=22,
            registration_sheet_sync_error="old error",
        )
        source_ticket = Ticket.objects.create(event=source, name="VIP", order=3)
        source_question = Question.objects.create(event=source, text="Why attend?", is_required=True, order=4)
        CheckIn.objects.create(event=source, name="Source entrance")
        original_count = Event.objects.count()

        response = self.client.get(f"/admin/event/event/add/?copy_from={source.pk}")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Event.objects.count(), original_count)
        self.assertContains(response, "Loaded safe template data from")
        self.assertContains(response, "<strong>Source Event</strong>", html=True)
        self.assertContains(response, "Source Event — May 14–16, 2026")
        self.assertContains(response, 'href="#tickets-group"')
        self.assertContains(response, 'href="#questions-group"')
        initial = response.context["adminform"].form.initial
        self.assertEqual(initial["name"], "")
        self.assertEqual(initial["slug"], "")
        self.assertEqual(str(initial["date"]), "2026-05-14")
        self.assertEqual(str(initial["end_date"]), "2026-05-16")
        self.assertEqual(initial["location"], "Source Hall")
        self.assertFalse(initial["registration_open"])
        self.assertTrue(initial["allow_secondary_email"])
        self.assertTrue(initial["collect_phone"])
        self.assertTrue(initial["verify_phone"])
        self.assertEqual(initial["ticket_login_validity_days"], 45)
        self.assertFalse(initial["ticket_login_reusable"])
        self.assertEqual(initial["registration_sheet_id"], "")
        self.assertIsNone(initial["registration_sheet_gid"])

        inline_forms = {inline.opts.model: inline.formset.forms for inline in response.context["inline_admin_formsets"]}
        self.assertEqual(inline_forms[Ticket][0].initial, {"name": "VIP", "order": 3})
        self.assertEqual(
            inline_forms[Question][0].initial,
            {"text": "Why attend?", "is_required": True, "order": 4},
        )
        source_ticket.refresh_from_db()
        source_question.refresh_from_db()

    def test_add_form_loads_multiple_ticket_types_and_questions_in_order(self):
        source = make_event(name="Multiple Children Source")
        Ticket.objects.create(event=source, name="VIP", order=30)
        Ticket.objects.create(event=source, name="General", order=10)
        Ticket.objects.create(event=source, name="Student", order=20)
        Question.objects.create(event=source, text="Optional note?", order=8)
        Question.objects.create(event=source, text="Required role?", is_required=True, order=2)

        response = self.client.get(f"/admin/event/event/add/?copy_from={source.pk}")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "3 Ticket types")
        self.assertContains(response, "2 Questions")
        inline_forms = {inline.opts.model: inline.formset.forms for inline in response.context["inline_admin_formsets"]}
        self.assertEqual(
            [form.initial for form in inline_forms[Ticket]],
            [
                {"name": "General", "order": 10},
                {"name": "Student", "order": 20},
                {"name": "VIP", "order": 30},
            ],
        )
        self.assertEqual(
            [form.initial for form in inline_forms[Question]],
            [
                {"text": "Required role?", "is_required": True, "order": 2},
                {"text": "Optional note?", "is_required": False, "order": 8},
            ],
        )

    def test_add_form_loads_source_with_zero_ticket_types_and_questions(self):
        source = make_event(name="Empty Children Source")

        response = self.client.get(f"/admin/event/event/add/?copy_from={source.pk}")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "0 Ticket types")
        self.assertContains(response, "0 Questions")
        inline_forms = {inline.opts.model: inline.formset.forms for inline in response.context["inline_admin_formsets"]}
        self.assertEqual(inline_forms[Ticket], [])
        self.assertEqual(inline_forms[Question], [])

    def test_add_form_rejects_missing_copy_source(self):
        deleted_source = make_event(name="Deleted Copy Source")
        deleted_source_id = deleted_source.pk
        deleted_source.delete()

        for source_id in ("not-a-valid-id", str(uuid.uuid4()), str(deleted_source_id)):
            with self.subTest(source_id=source_id):
                response = self.client.get(f"/admin/event/event/add/?copy_from={source_id}")
                self.assertEqual(response.status_code, 404)

    @override_settings(ADMIN_REQUIRE_CONFIRMATION=False)
    def test_saving_loaded_copy_creates_new_children_and_resets_operational_data(self):
        source = make_event(
            name="Save Source Event",
            registration_open=True,
            allow_secondary_email=True,
            collect_phone=True,
            verify_phone=True,
            registration_sheet_id="source-sheet",
            registration_sheet_gid=42,
        )
        source_ticket = Ticket.objects.create(event=source, name="General Admission", order=1)
        source_question = Question.objects.create(event=source, text="Role?", is_required=True, order=2)
        CheckIn.objects.create(event=source, name="Source entrance")

        response = self.client.post(
            f"/admin/event/event/add/?copy_from={source.pk}",
            {
                "name": "New Copied Event",
                "slug": "new-copied-event",
                "date": "2026-10-01",
                "end_date": "2026-10-03",
                "location": source.location,
                "description": source.description,
                "allow_secondary_email": "on",
                "collect_phone": "on",
                "verify_phone": "on",
                "ticket_login_validity_days": "30",
                "ticket_login_reusable": "on",
                "registration_sheet_id": "",
                "registration_sheet_gid": "",
                "tickets-TOTAL_FORMS": "1",
                "tickets-INITIAL_FORMS": "0",
                "tickets-MIN_NUM_FORMS": "0",
                "tickets-MAX_NUM_FORMS": "1000",
                "tickets-0-name": source_ticket.name,
                "tickets-0-order": str(source_ticket.order),
                "questions-TOTAL_FORMS": "1",
                "questions-INITIAL_FORMS": "0",
                "questions-MIN_NUM_FORMS": "0",
                "questions-MAX_NUM_FORMS": "1000",
                "questions-0-text": source_question.text,
                "questions-0-is_required": "on",
                "questions-0-order": str(source_question.order),
                "_save": "Save",
            },
        )

        self.assertEqual(response.status_code, 302)
        copied = Event.objects.get(slug="new-copied-event")
        self.assertFalse(copied.registration_open)
        self.assertTrue(copied.allow_secondary_email)
        self.assertTrue(copied.collect_phone)
        self.assertTrue(copied.verify_phone)
        self.assertEqual(copied.ticket_login_validity_days, source.ticket_login_validity_days)
        self.assertEqual(copied.ticket_login_reusable, source.ticket_login_reusable)
        self.assertEqual(copied.registration_sheet_id, "")
        self.assertIsNone(copied.registration_sheet_gid)
        self.assertEqual(copied.registration_sheet_sync_count, 0)
        self.assertEqual(copied.registration_sheet_sync_error, "")
        self.assertEqual(copied.check_ins.count(), 0)
        copied_ticket = copied.tickets.get(name=source_ticket.name)
        copied_question = copied.questions.get(text=source_question.text)
        self.assertNotEqual(copied_ticket.pk, source_ticket.pk)
        self.assertNotEqual(copied_ticket.barcode, source_ticket.barcode)
        self.assertNotEqual(copied_question.pk, source_question.pk)
        source.refresh_from_db()
        self.assertTrue(source.registration_open)
        self.assertEqual(source.registration_sheet_id, "source-sheet")

    @override_settings(ADMIN_REQUIRE_CONFIRMATION=False)
    def test_saving_loaded_copy_uses_final_edited_deleted_and_added_inline_rows(self):
        source = make_event(name="Editable Copy Source", registration_open=True)
        source_tickets = [
            Ticket.objects.create(event=source, name="General", order=1),
            Ticket.objects.create(event=source, name="VIP", order=2),
            Ticket.objects.create(event=source, name="Student", order=3),
        ]
        source_questions = [
            Question.objects.create(event=source, text="Role?", is_required=True, order=1),
            Question.objects.create(event=source, text="Dietary needs?", order=2),
        ]
        source_ticket_snapshot = list(source.tickets.values("pk", "name", "order", "barcode"))
        source_question_snapshot = list(source.questions.values("pk", "text", "is_required", "order"))
        self.client.get(f"/admin/event/event/add/?copy_from={source.pk}")

        response = self.client.post(
            f"/admin/event/event/add/?copy_from={source.pk}",
            self._copy_post_data(
                name="Reviewed Copy",
                slug="reviewed-copy",
                ticket_rows=[
                    {"name": "General Admission", "order": 10},
                    {"name": "VIP", "order": 2, "DELETE": True},
                    {"name": "Student", "order": 3},
                    {"name": "Sponsor", "order": 4},
                ],
                question_rows=[
                    {"text": "Your role?", "is_required": False, "order": 5},
                    {"text": "Dietary needs?", "order": 2, "DELETE": True},
                    {"text": "Accessibility needs?", "is_required": True, "order": 3},
                ],
            ),
        )

        self.assertEqual(response.status_code, 302)
        copied = Event.objects.get(slug="reviewed-copy")
        self.assertEqual(
            list(copied.tickets.values_list("name", "order")),
            [("Student", 3), ("Sponsor", 4), ("General Admission", 10)],
        )
        self.assertEqual(
            list(copied.questions.values_list("text", "is_required", "order")),
            [("Accessibility needs?", True, 3), ("Your role?", False, 5)],
        )
        source_ticket_ids = {ticket.pk for ticket in source_tickets}
        source_ticket_barcodes = {ticket.barcode for ticket in source_tickets}
        source_question_ids = {question.pk for question in source_questions}
        self.assertTrue(source_ticket_ids.isdisjoint(copied.tickets.values_list("pk", flat=True)))
        self.assertTrue(source_ticket_barcodes.isdisjoint(copied.tickets.values_list("barcode", flat=True)))
        self.assertTrue(source_question_ids.isdisjoint(copied.questions.values_list("pk", flat=True)))
        source.refresh_from_db()
        self.assertTrue(source.registration_open)
        self.assertEqual(list(source.tickets.values("pk", "name", "order", "barcode")), source_ticket_snapshot)
        self.assertEqual(
            list(source.questions.values("pk", "text", "is_required", "order")),
            source_question_snapshot,
        )

    @override_settings(ADMIN_REQUIRE_CONFIRMATION=True)
    def test_post_uses_loaded_snapshot_when_source_was_deleted_after_get(self):
        source = make_event(name="Ephemeral Copy Source")
        source_id = source.pk
        Ticket.objects.create(event=source, name="Snapshot Ticket", order=2)
        Question.objects.create(event=source, text="Snapshot question?", is_required=True, order=3)
        self.assertEqual(self.client.get(f"/admin/event/event/add/?copy_from={source_id}").status_code, 200)
        source.delete()

        response = self.client.post(
            f"/admin/event/event/add/?copy_from={source_id}",
            self._copy_post_data(
                name="Saved Detached Snapshot",
                slug="saved-detached-snapshot",
                ticket_rows=[{"name": "Snapshot Ticket", "order": 2}],
                question_rows=[{"text": "Snapshot question?", "is_required": True, "order": 3}],
            ),
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn("confirm-change", response.url)
        pending = self.client.session["_admin_pending_change_event_event"]
        response = self.client.post(
            response.url,
            {"token": pending["token"], "confirmation_word": "event"},
        )
        self.assertEqual(response.status_code, 302)
        copied = Event.objects.get(slug="saved-detached-snapshot")
        self.assertEqual(list(copied.tickets.values_list("name", flat=True)), ["Snapshot Ticket"])
        self.assertEqual(list(copied.questions.values_list("text", flat=True)), ["Snapshot question?"])
        self.assertFalse(Event.objects.filter(pk=source_id).exists())

    @override_settings(ADMIN_REQUIRE_CONFIRMATION=True)
    def test_confirmation_reviews_final_children_then_saves_new_rows_without_changing_source(self):
        source = make_event(name="Confirmation Copy Source", registration_open=True)
        kept_ticket = Ticket.objects.create(event=source, name="General", order=1)
        deleted_ticket = Ticket.objects.create(event=source, name="Remove VIP", order=2)
        kept_question = Question.objects.create(event=source, text="Role?", is_required=True, order=1)
        deleted_question = Question.objects.create(event=source, text="Remove question?", order=2)
        original_ticket_rows = list(source.tickets.values("pk", "name", "order", "barcode"))
        original_question_rows = list(source.questions.values("pk", "text", "is_required", "order"))
        self.client.get(f"/admin/event/event/add/?copy_from={source.pk}")

        response = self.client.post(
            f"/admin/event/event/add/?copy_from={source.pk}",
            self._copy_post_data(
                name="Confirmed Event Copy",
                slug="confirmed-event-copy",
                ticket_rows=[
                    {"name": "General Edited", "order": 5},
                    {"name": deleted_ticket.name, "order": deleted_ticket.order, "DELETE": True},
                    {"name": "New Sponsor", "order": 2},
                ],
                question_rows=[
                    {"text": "Role edited?", "is_required": False, "order": 4},
                    {"text": deleted_question.text, "order": deleted_question.order, "DELETE": True},
                    {"text": "New required question?", "is_required": True, "order": 3},
                ],
            ),
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn("confirm-change", response.url)
        self.assertFalse(Event.objects.filter(slug="confirmed-event-copy").exists())

        confirmation = self.client.get(response.url)
        self.assertEqual(confirmation.status_code, 200)
        child_diff = {
            item["label"]: item["new_value"]
            for item in confirmation.context["diff"]
            if item["field"].startswith(("tickets", "questions"))
        }
        self.assertEqual(
            child_diff,
            {
                "Ticket type 1": "New Sponsor — order 2",
                "Ticket type 2": "General Edited — order 5",
                "Question 1": "New required question? — required: Yes; order 3",
                "Question 2": "Role edited? — required: No; order 4",
            },
        )
        self.assertNotContains(confirmation, deleted_ticket.name)
        self.assertNotContains(confirmation, deleted_question.text)

        pending = self.client.session["_admin_pending_change_event_event"]
        confirmed = self.client.post(
            response.url,
            {"token": pending["token"], "confirmation_word": "event"},
        )

        self.assertEqual(confirmed.status_code, 302)
        copied = Event.objects.get(slug="confirmed-event-copy")
        self.assertEqual(
            list(copied.tickets.values_list("name", "order")),
            [("New Sponsor", 2), ("General Edited", 5)],
        )
        self.assertEqual(
            list(copied.questions.values_list("text", "is_required", "order")),
            [("New required question?", True, 3), ("Role edited?", False, 4)],
        )
        self.assertNotIn(kept_ticket.pk, copied.tickets.values_list("pk", flat=True))
        self.assertNotIn(kept_ticket.barcode, copied.tickets.values_list("barcode", flat=True))
        self.assertNotIn(kept_question.pk, copied.questions.values_list("pk", flat=True))
        source.refresh_from_db()
        self.assertTrue(source.registration_open)
        self.assertEqual(list(source.tickets.values("pk", "name", "order", "barcode")), original_ticket_rows)
        self.assertEqual(
            list(source.questions.values("pk", "text", "is_required", "order")),
            original_question_rows,
        )

    @override_settings(ADMIN_REQUIRE_CONFIRMATION=True)
    def test_confirmation_explicitly_shows_none_for_zero_children(self):
        response = self.client.post(
            "/admin/event/event/add/",
            self._copy_post_data(name="No Children Copy", slug="no-children-copy"),
        )

        self.assertEqual(response.status_code, 302)
        confirmation = self.client.get(response.url)
        child_diff = {
            item["label"]: item["new_value"]
            for item in confirmation.context["diff"]
            if item["field"] in {"tickets", "questions"}
        }
        self.assertEqual(child_diff, {"Ticket types": "None", "Questions": "None"})

    @override_settings(ADMIN_REQUIRE_CONFIRMATION=True)
    def test_invalid_children_stay_on_add_form_without_pending_confirmation_or_event(self):
        response = self.client.post(
            "/admin/event/event/add/",
            self._copy_post_data(
                name="Invalid Children Copy",
                slug="invalid-children-copy",
                ticket_rows=[{"name": "", "order": 1}],
                question_rows=[{"text": "", "is_required": True, "order": 1}],
            ),
        )

        self.assertEqual(response.status_code, 200)
        inline_formsets = {inline.opts.model: inline.formset for inline in response.context["inline_admin_formsets"]}
        self.assertIn("name", inline_formsets[Ticket].forms[0].errors)
        self.assertIn("text", inline_formsets[Question].forms[0].errors)
        self.assertNotIn("_admin_pending_change_event_event", self.client.session)
        self.assertFalse(Event.objects.filter(slug="invalid-children-copy").exists())

    def test_change_page_shows_inlines_for_staff_with_event_app_access(self):
        """Inlines match Event admin access (the ``event`` app grant), not per-model
        Django permissions — see apps.core.access.user_can_access_app."""
        editor = Member.objects.create_user(
            password="testpass123", is_staff=True, is_superuser=False, admin_apps=["event"]
        )
        ContactEmail.objects.create(
            member=editor,
            email_address="editor@example.com",
            email_type="primary",
            verified=True,
        )
        event = make_event(name="Inline Perm Test")
        self.client.logout()
        self.client.login(username="editor@example.com", password="testpass123")

        response = self.client.get(f"/admin/event/event/{event.pk}/change/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="tickets-group"')
        self.assertContains(response, 'id="questions-group"')

    @override_settings(ADMIN_REQUIRE_CONFIRMATION=True)
    def test_event_only_staff_can_load_and_save_copied_children(self):
        source = make_event(name="Staff Copy Source")
        Ticket.objects.create(event=source, name="Staff Ticket", order=1)
        Question.objects.create(event=source, text="Staff question?", is_required=True, order=2)
        editor = Member.objects.create_user(
            password="testpass123",
            is_staff=True,
            is_superuser=False,
            admin_apps=["event"],
        )
        ContactEmail.objects.create(
            member=editor,
            email_address="copy-editor@example.com",
            email_type="primary",
            verified=True,
        )
        self.client.logout()
        self.client.login(username="copy-editor@example.com", password="testpass123")

        loaded = self.client.get(f"/admin/event/event/add/?copy_from={source.pk}")

        self.assertEqual(loaded.status_code, 200)
        self.assertContains(loaded, 'value="Staff Ticket"')
        self.assertContains(loaded, 'value="Staff question?"')
        saved = self.client.post(
            f"/admin/event/event/add/?copy_from={source.pk}",
            self._copy_post_data(
                name="Staff Saved Copy",
                slug="staff-saved-copy",
                ticket_rows=[{"name": "Staff Ticket", "order": 1}],
                question_rows=[{"text": "Staff question?", "is_required": True, "order": 2}],
            ),
        )
        self.assertEqual(saved.status_code, 302)
        self.assertIn("confirm-change", saved.url)
        self.assertEqual(self.client.get(saved.url).status_code, 200)
        pending = self.client.session["_admin_pending_change_event_event"]
        saved = self.client.post(
            saved.url,
            {"token": pending["token"], "confirmation_word": "event"},
        )
        self.assertEqual(saved.status_code, 302)
        copied = Event.objects.get(slug="staff-saved-copy")
        self.assertEqual(copied.tickets.get().name, "Staff Ticket")
        self.assertEqual(copied.questions.get().text, "Staff question?")

    def test_search_by_name(self):
        make_event(name="Searchable Event")
        response = self.client.get("/admin/event/event/?q=Searchable")
        self.assertEqual(response.status_code, 200)

    def test_list_filter_by_registration_open(self):
        response = self.client.get("/admin/event/event/?registration_open__exact=1")
        self.assertEqual(response.status_code, 200)

    @patch("apps.event.admin.current_project.sync_schedule")
    def test_pull_schedule_action_triggers_sync(self, mock_sync):
        from apps.event.models import CurrentProjectSchedule

        mock_sync.return_value = ScheduleSyncStats(sections_created=3, tracks_created=3, slots_created=4)
        config = CurrentProjectSchedule.objects.create(name="Demo Day")

        response = self.client.post("/admin/event/currentprojectschedule/pull/")

        self.assertEqual(response.status_code, 302)
        mock_sync.assert_called_once_with(config, sync_type="manual")


class CheckInAdminTest(TestCase):
    def setUp(self):
        self.admin_user = make_superuser(email="checkin-admin@example.com")
        self.client.login(username="checkin-admin@example.com", password="testpass123")
        self.event = make_event(name="Admin Check-in Event")
        self.check_in = CheckIn.objects.create(event=self.event, name="Main Entrance")

    def test_change_page_shows_live_summary_panel(self):
        response = self.client.get(f"/admin/event/checkin/{self.check_in.pk}/change/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "checkin-summary-config")
        self.assertContains(response, "event/css/checkin_change_summary.css")
        self.assertContains(response, "event/js/checkin_change_summary.js")
        self.assertContains(response, "data-checkin-summary")
        self.assertContains(response, "data-summary-total")
        self.assertContains(response, "data-summary-recent-list")
        self.assertContains(response, "data-summary-lookup-search")
        self.assertContains(response, "data-summary-ticket-filter")
        self.assertContains(response, "data-summary-lookup-list")
        self.assertContains(response, "Registration lookup")
        self.assertContains(response, "All ticket types")
        self.assertContains(response, "Ticket Type")
        self.assertContains(response, "Name, email, organization, title, code")
        self.assertContains(response, "Loading recent scans")
        self.assertContains(response, "pollIntervalMs")
        self.assertContains(response, f"/event/check-in/{self.check_in.pk}/status/")
        self.assertContains(response, "Open Check-in Console")
        self.assertContains(response, "Export Excel")
        self.assertContains(response, reverse("admin:event_checkin_scanner", args=[self.check_in.pk]))
        self.assertContains(response, reverse("admin:event_checkin_export", args=[self.check_in.pk]))
        self.assertNotContains(response, "This station")
        self.assertNotContains(response, "data-summary-station")
        self.assertNotContains(response, "Last 5 at this station")
        self.assertNotContains(response, "VIP Gate")
        self.assertNotContains(response, "Check in records")

    def test_changelist_hides_station_scan_count_column(self):
        ticket = make_ticket(self.event, name="General")
        attendee = make_member(email="list-checked-in@example.com", first_name="Ada", last_name="Lovelace")
        registration = make_registration(attendee, self.event, ticket)
        CheckInRecord.objects.create(check_in=self.check_in, registration=registration)

        response = self.client.get("/admin/event/checkin/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Open Console")
        self.assertNotContains(response, "Scans")

    def test_export_excel_includes_checkin_status_and_member_profile(self):
        ticket = make_ticket(self.event, name="VIP")
        member = make_member(
            email="checkin-export@example.com",
            first_name="Ada",
            last_name="Lovelace",
            organization="Analytical Engines",
            title="Chief Scientist",
        )
        registration = make_registration(
            member,
            self.event,
            ticket,
            attendee_organization="Registration Org",
        )
        CheckInRecord.objects.create(check_in=self.check_in, registration=registration, scanned_by=self.admin_user)

        response = self.client.get(reverse("admin:event_checkin_export", args=[self.check_in.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("checkin_admin-check-in-event_main-entrance_", response["Content-Disposition"])
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(
            rows[0],
            (
                "Event",
                "Check-in",
                "Ticket Type",
                "Ticket Code",
                "Attendee Name",
                "Attendee Email",
                "Attendee Organization",
                "Member Title",
                "Member Organization",
                "Checked In",
                "Check-in Station",
                "Checked-in At",
                "Scanned By",
            ),
        )
        self.assertEqual(
            rows[1][0:11],
            (
                "Admin Check-in Event",
                "Main Entrance",
                "VIP",
                registration.ticket_code,
                "Ada Lovelace",
                "checkin-export@example.com",
                "Registration Org",
                "Chief Scientist",
                "Analytical Engines",
                "Yes",
                "Main Entrance",
            ),
        )

    def test_change_page_live_summary_hides_station_count_when_scans_exist(self):
        ticket = make_ticket(self.event, name="General")
        attendee = make_member(email="checked-in@example.com", first_name="Ada", last_name="Lovelace")
        registration = make_registration(attendee, self.event, ticket)
        CheckInRecord.objects.create(check_in=self.check_in, registration=registration)

        response = self.client.get(f"/admin/event/checkin/{self.check_in.pk}/change/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "data-summary-scanned")
        self.assertNotContains(response, "This station")
        self.assertNotContains(response, "data-summary-station")

    def test_scanner_page_is_available_to_staff(self):
        response = self.client.get(reverse("admin:event_checkin_scanner", args=[self.check_in.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "checkin-console-config")
        self.assertContains(response, "event/css/checkin_console.css")
        self.assertContains(response, "event/js/checkin_console.js")
        self.assertContains(response, "data-sync-status")
        self.assertContains(response, "statusPollIntervalMs")
        self.assertContains(response, "data-camera-message")
        self.assertContains(response, "Request camera access")
        self.assertNotContains(response, "This station")
        self.assertNotContains(response, "data-stat-station")
        self.assertContains(response, f"/event/check-in/{self.check_in.pk}/scan/")
        self.assertContains(response, f"/event/check-in/{self.check_in.pk}/status/")
        self.assertContains(response, f"/event/check-in/{self.check_in.pk}/records/__record_id__/undo/")
        self.assertNotContains(response, "checkin-state")
        self.assertNotContains(response, "function startCamera")

    def test_scanner_page_rejects_non_staff_user(self):
        user = make_member(email="nonstaff-checkin@example.com")
        self.client.force_login(user)

        response = self.client.get(reverse("admin:event_checkin_scanner", args=[self.check_in.pk]))

        self.assertNotEqual(response.status_code, 200)


@override_settings(ADMIN_REQUIRE_CONFIRMATION=False)
class EventRegistrationAdminTest(TestCase):
    def setUp(self):
        self.admin_user = make_superuser()
        self.client.login(username="admin@example.com", password="testpass123")

    def test_changelist_accessible(self):
        response = self.client.get("/admin/event/eventregistration/")
        self.assertEqual(response.status_code, 200)

    def test_changelist_shows_send_all_ticket_emails_button(self):
        response = self.client.get("/admin/event/eventregistration/")
        self.assertContains(response, "Send All Tickets")
        self.assertContains(response, reverse("admin:event_eventregistration_send_all_ticket_emails"))

    def test_export_column_picker_includes_member_information(self):
        event = make_event()
        ticket = make_ticket(event)
        member = make_member(email="export-picker@example.com")
        registration = make_registration(member, event, ticket)

        response = self.client.post(
            "/admin/event/eventregistration/",
            {
                "action": "export_data",
                ACTION_CHECKBOX_NAME: str(registration.pk),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Member Title")
        self.assertContains(response, "Member Organization")
        self.assertContains(response, "Member Primary Email")
        self.assertContains(response, "Member Phone Numbers")
        self.assertContains(response, "Event Start Date")
        self.assertContains(response, "Event End Date")
        self.assertNotContains(response, 'value="event_date"')

    def test_export_includes_event_start_and_end_dates(self):
        event = make_event(name="Multi-day Export", date="2025-06-15", end_date="2025-06-17")
        ticket = make_ticket(event)
        member = make_member(email="export-dates@example.com")
        registration = make_registration(member, event, ticket)

        response = self.client.post(
            "/admin/event/eventregistration/",
            {
                "action": "export_data",
                ACTION_CHECKBOX_NAME: str(registration.pk),
                "export_confirm": "1",
                "export_format": "json",
                "export_filename": "registration_dates",
                "export_fields": ["event_date", "event_start_date", "event_end_date"],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            json.loads(response.content),
            [
                {
                    "event_date": "2025-06-15",
                    "event_start_date": "2025-06-15",
                    "event_end_date": "2025-06-17",
                }
            ],
        )

    def test_export_accepts_legacy_event_date_selection(self):
        event = make_event(name="Legacy Export", date="2025-06-15", end_date="2025-06-17")
        ticket = make_ticket(event)
        member = make_member(email="legacy-export-date@example.com")
        registration = make_registration(member, event, ticket)

        response = self.client.post(
            "/admin/event/eventregistration/",
            {
                "action": "export_data",
                ACTION_CHECKBOX_NAME: str(registration.pk),
                "export_confirm": "1",
                "export_format": "json",
                "export_fields": ["event_date"],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(json.loads(response.content), [{"event_date": "2025-06-15"}])

    def test_export_old_null_end_date_falls_back_to_start_date(self):
        event = make_event(name="Rolling Deploy Export", date="2025-06-15")
        Event.objects.filter(pk=event.pk).update(end_date=None)
        ticket = make_ticket(event)
        member = make_member(email="export-null-end-date@example.com")
        registration = make_registration(member, event, ticket)

        response = self.client.post(
            "/admin/event/eventregistration/",
            {
                "action": "export_data",
                ACTION_CHECKBOX_NAME: str(registration.pk),
                "export_confirm": "1",
                "export_format": "json",
                "export_filename": "registration_legacy_date",
                "export_fields": ["event_date", "event_start_date", "event_end_date"],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            json.loads(response.content),
            [
                {
                    "event_date": "2025-06-15",
                    "event_start_date": "2025-06-15",
                    "event_end_date": "2025-06-15",
                }
            ],
        )

    def test_export_excel_includes_member_profile_and_contacts(self):
        event = make_event(name="Export Event")
        ticket = make_ticket(event, name="VIP")
        member = make_member(
            email="export-primary@example.com",
            first_name="Ada",
            middle_name="Byron",
            last_name="Lovelace",
            organization="Analytical Engines",
            title="Chief Scientist",
        )
        ContactEmail.objects.create(
            member=member,
            email_address="export-secondary@example.com",
            email_type="secondary",
            verified=True,
        )
        ContactEmail.objects.create(
            member=member,
            email_address="export-other@example.com",
            email_type="other",
            verified=False,
        )
        ContactPhone.objects.create(
            member=member,
            phone_number="2095551212",
            region="1-US",
            verified=True,
        )
        registration = make_registration(
            member,
            event,
            ticket,
            attendee_first_name="Grace",
            attendee_last_name="Hopper",
            attendee_email="grace@example.com",
            attendee_secondary_email="grace.secondary@example.com",
            attendee_phone="2095552323",
            attendee_organization="Navy",
        )

        response = self.client.post(
            "/admin/event/eventregistration/",
            {
                "action": "export_data",
                ACTION_CHECKBOX_NAME: str(registration.pk),
                "export_confirm": "1",
                "export_format": "xlsx",
                "export_filename": "registration_export",
                "export_fields": [
                    "event_name",
                    "ticket_name",
                    "attendee_email",
                    "attendee_organization",
                    "member_full_name",
                    "member_title",
                    "member_organization",
                    "member_primary_email",
                    "member_secondary_emails",
                    "member_other_emails",
                    "member_phone_numbers",
                ],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(
            rows[0],
            (
                "Event Name",
                "Ticket",
                "Attendee Email",
                "Attendee Organization",
                "Member Full Name",
                "Member Title",
                "Member Organization",
                "Member Primary Email",
                "Member Secondary Emails",
                "Member Other Emails",
                "Member Phone Numbers",
            ),
        )
        self.assertEqual(
            rows[1],
            (
                "Export Event",
                "VIP",
                "grace@example.com",
                "Navy",
                "Ada Byron Lovelace",
                "Chief Scientist",
                "Analytical Engines",
                "export-primary@example.com",
                "export-secondary@example.com",
                "export-other@example.com",
                "(209)555-1212",
            ),
        )

    def test_send_all_ticket_emails_confirmation_page(self):
        event = make_event()
        ticket = make_ticket(event)
        member = make_member(email="ticket-confirm@example.com")
        make_registration(member, event, ticket)

        response = self.client.get(reverse("admin:event_eventregistration_send_all_ticket_emails"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Send Ticket Emails to All Registrants")
        self.assertContains(response, "You are about to send ticket emails to 1 registrant")

    @patch("apps.event.services.ticket_mail.send_ticket_email")
    def test_send_all_ticket_emails_posts_all_registrations(self, mock_send):
        event = make_event()
        ticket = make_ticket(event)
        member_one = make_member(email="ticket-one@example.com")
        member_two = make_member(email="ticket-two@example.com")
        registration_one = make_registration(member_one, event, ticket)
        registration_two = make_registration(member_two, event, ticket)

        response = self.client.post(reverse("admin:event_eventregistration_send_all_ticket_emails"))

        self.assertRedirects(response, reverse("admin:event_eventregistration_changelist"))
        self.assertEqual(mock_send.call_count, 2)
        self.assertEqual(
            {call.args[0].pk for call in mock_send.call_args_list},
            {registration_one.pk, registration_two.pk},
        )

    @patch("apps.event.services.ticket_mail.send_ticket_email")
    def test_send_all_ticket_emails_empty_queryset_does_not_send(self, mock_send):
        response = self.client.post(reverse("admin:event_eventregistration_send_all_ticket_emails"))

        self.assertRedirects(response, reverse("admin:event_eventregistration_changelist"))
        mock_send.assert_not_called()

    def test_has_add_permission(self):
        from django.test import RequestFactory

        factory = RequestFactory()
        request = factory.get("/admin/")
        request.user = self.admin_user
        admin_instance = EventRegistrationAdmin(EventRegistration, None)
        self.assertTrue(admin_instance.has_add_permission(request))

    def test_add_form_accessible(self):
        response = self.client.get("/admin/event/eventregistration/add/")
        self.assertEqual(response.status_code, 200)

    @patch("apps.event.services.registration_sheet_sync.schedule_registration_sync")
    def test_admin_add_creates_registration(self, mock_sync):
        member = Member.objects.create_user(password="testpass123")
        ContactEmail.objects.create(
            member=member,
            email_address="reg@e.com",
            email_type="primary",
            verified=True,
        )
        event = make_event()
        ticket = Ticket.objects.create(event=event, name="GA")
        response = self.client.post(
            "/admin/event/eventregistration/add/",
            {
                "member": str(member.pk),
                "event": str(event.pk),
                "ticket": str(ticket.pk),
                "attendee_first_name": "",
                "attendee_last_name": "",
                "attendee_email": "",
                "attendee_secondary_email": "",
                "attendee_phone": "",
                "attendee_organization": "",
                "phone_verified": "",
                "question_answers": "[]",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(EventRegistration.objects.filter(member=member, event=event).exists())
        mock_sync.assert_called_once_with(event)

    def test_admin_add_rejects_duplicate(self):
        member = Member.objects.create_user(password="testpass123")
        ContactEmail.objects.create(
            member=member,
            email_address="dup@e.com",
            email_type="primary",
            verified=True,
        )
        event = make_event()
        ticket = Ticket.objects.create(event=event, name="GA")
        EventRegistration.objects.create(member=member, event=event, ticket=ticket)
        response = self.client.post(
            "/admin/event/eventregistration/add/",
            {
                "member": str(member.pk),
                "event": str(event.pk),
                "ticket": str(ticket.pk),
                "attendee_first_name": "",
                "attendee_last_name": "",
                "attendee_email": "",
                "attendee_secondary_email": "",
                "attendee_phone": "",
                "attendee_organization": "",
                "phone_verified": "",
                "question_answers": "[]",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(EventRegistration.objects.filter(member=member, event=event).count(), 1)

    def test_admin_add_rejects_mismatched_ticket(self):
        member = Member.objects.create_user(password="testpass123")
        ContactEmail.objects.create(
            member=member,
            email_address="mis@e.com",
            email_type="primary",
            verified=True,
        )
        event1 = make_event(name="Event 1")
        event2 = make_event(name="Event 2")
        ticket_from_event2 = Ticket.objects.create(event=event2, name="VIP")
        response = self.client.post(
            "/admin/event/eventregistration/add/",
            {
                "member": str(member.pk),
                "event": str(event1.pk),
                "ticket": str(ticket_from_event2.pk),
                "attendee_first_name": "",
                "attendee_last_name": "",
                "attendee_email": "",
                "attendee_secondary_email": "",
                "attendee_phone": "",
                "attendee_organization": "",
                "phone_verified": "",
                "question_answers": "[]",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(EventRegistration.objects.filter(member=member, event=event1).exists())

    def test_has_change_permission(self):
        from django.test import RequestFactory

        factory = RequestFactory()
        request = factory.get("/admin/")
        request.user = self.admin_user
        admin_instance = EventRegistrationAdmin(EventRegistration, None)
        self.assertTrue(admin_instance.has_change_permission(request))

    def test_change_page_keeps_only_ticket_editable_in_ticket_section(self):
        from django.test import RequestFactory

        event = make_event()
        ticket = make_ticket(event)
        member = make_member(email="readonly-ticket-section@example.com")
        registration = make_registration(member, event, ticket)
        factory = RequestFactory()
        request = factory.get("/admin/")
        request.user = self.admin_user
        admin_instance = EventRegistrationAdmin(EventRegistration, None)

        readonly_fields = admin_instance.get_readonly_fields(request, obj=registration)

        self.assertNotIn("ticket", readonly_fields)
        self.assertIn("event", readonly_fields)
        self.assertIn("member", readonly_fields)
        self.assertIn("ticket_code", readonly_fields)
        self.assertIn("attendee_email", readonly_fields)
        self.assertIn("send_ticket_email_action", readonly_fields)

    def test_change_form_limits_ticket_choices_to_registration_event(self):
        event = make_event(name="Ticket Choice Event")
        current_ticket = make_ticket(event, name="Current Ticket")
        new_ticket = make_ticket(event, name="Upgrade Ticket")
        other_event = make_event(name="Other Ticket Event")
        other_ticket = make_ticket(other_event, name="Other Event Ticket")
        member = make_member(email="ticket-choice@example.com")
        registration = make_registration(member, event, current_ticket)

        response = self.client.get(f"/admin/event/eventregistration/{registration.pk}/change/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="ticket"')
        self.assertContains(response, 'name="_send_ticket_email"')
        self.assertContains(response, "Send ticket email now")
        self.assertNotContains(response, 'name="send_ticket_email"')
        self.assertNotContains(response, "If checked, a confirmation email")
        self.assertContains(response, str(new_ticket))
        self.assertNotContains(response, str(other_ticket))

    @patch("apps.event.services.registration_sheet_sync.schedule_registration_sync")
    def test_admin_change_updates_registration_ticket(self, mock_sync):
        event = make_event()
        current_ticket = make_ticket(event, name="General")
        new_ticket = make_ticket(event, name="VIP")
        member = make_member(email="ticket-change@example.com")
        registration = make_registration(member, event, current_ticket)

        response = self.client.post(
            f"/admin/event/eventregistration/{registration.pk}/change/",
            {
                "ticket": str(new_ticket.pk),
                "_save": "Save",
            },
        )

        self.assertEqual(response.status_code, 302)
        registration.refresh_from_db()
        self.assertEqual(registration.ticket, new_ticket)
        mock_sync.assert_called_once_with(event)

    @patch("apps.event.services.registration_sheet_sync.schedule_registration_sync")
    def test_admin_change_rejects_ticket_from_other_event(self, mock_sync):
        event = make_event(name="Original Event")
        current_ticket = make_ticket(event, name="General")
        other_event = make_event(name="Other Event")
        other_ticket = make_ticket(other_event, name="Other VIP")
        member = make_member(email="ticket-mismatch-change@example.com")
        registration = make_registration(member, event, current_ticket)

        response = self.client.post(
            f"/admin/event/eventregistration/{registration.pk}/change/",
            {
                "ticket": str(other_ticket.pk),
                "_save": "Save",
            },
        )

        self.assertEqual(response.status_code, 200)
        registration.refresh_from_db()
        self.assertEqual(registration.ticket, current_ticket)
        mock_sync.assert_not_called()

    @patch("apps.event.services.ticket_mail.send_ticket_email")
    @patch("apps.event.services.registration_sheet_sync.schedule_registration_sync")
    def test_admin_change_send_ticket_button_saves_ticket_and_sends_email(self, mock_sync, mock_send):
        event = make_event()
        current_ticket = make_ticket(event, name="General")
        new_ticket = make_ticket(event, name="VIP")
        member = make_member(email="ticket-send-button@example.com")
        registration = make_registration(member, event, current_ticket)

        response = self.client.post(
            f"/admin/event/eventregistration/{registration.pk}/change/",
            {
                "ticket": str(new_ticket.pk),
                "_send_ticket_email": "1",
            },
        )

        self.assertRedirects(response, f"/admin/event/eventregistration/{registration.pk}/change/")
        registration.refresh_from_db()
        self.assertEqual(registration.ticket, new_ticket)
        mock_sync.assert_called_once_with(event)
        mock_send.assert_called_once()
        self.assertEqual(mock_send.call_args.args[0].pk, registration.pk)
        self.assertEqual(mock_send.call_args.args[0].ticket_id, new_ticket.pk)

    def test_has_delete_permission(self):
        from django.test import RequestFactory

        factory = RequestFactory()
        request = factory.get("/admin/")
        request.user = self.admin_user
        admin_instance = EventRegistrationAdmin(EventRegistration, None)
        self.assertTrue(admin_instance.has_delete_permission(request))

    def test_search_by_attendee_fields(self):
        member = Member.objects.create_user(password="testpass123")
        ContactEmail.objects.create(member=member, email_address="u@e.com", email_type="primary", verified=True)
        event = make_event()
        ticket = Ticket.objects.create(event=event, name="GA")
        EventRegistration.objects.create(
            member=member,
            event=event,
            ticket=ticket,
            attendee_first_name="Searchable",
            attendee_last_name="Name",
        )
        response = self.client.get("/admin/event/eventregistration/?q=Searchable")
        self.assertEqual(response.status_code, 200)

    def test_member_info_endpoint(self):
        member = Member.objects.create_user(
            password="testpass123",
            first_name="Ada",
            last_name="Lovelace",
            organization="UCM",
            title="Prof",
        )
        ContactEmail.objects.create(
            member=member,
            email_address="ada@e.com",
            email_type="primary",
            verified=True,
        )
        ContactEmail.objects.create(
            member=member,
            email_address="ada2@e.com",
            email_type="secondary",
            verified=False,
        )
        response = self.client.get(f"/admin/event/eventregistration/member-info/{member.pk}/")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["name"], "Ada Lovelace")
        self.assertIn("ada@e.com", data["emails"])
        self.assertIn("ada2@e.com", data["emails"])
        self.assertEqual(data["organization"], "UCM")
        self.assertEqual(data["title"], "Prof")

    def test_event_info_endpoint(self):
        event = make_event(
            name="Info Test",
            end_date="2025-06-17",
            registration_open=True,
            collect_phone=True,
            verify_phone=True,
        )
        Ticket.objects.create(event=event, name="GA")
        Ticket.objects.create(event=event, name="VIP")
        response = self.client.get(f"/admin/event/eventregistration/event-info/{event.pk}/")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["name"], "Info Test")
        self.assertEqual(data["date"], "2025-06-15")
        self.assertEqual(data["end_date"], "2025-06-17")
        self.assertEqual(data["date_range"], "June 15–17, 2025")
        self.assertEqual(data["location"], "Test Venue")
        self.assertNotIn("is_live", data)
        self.assertTrue(data["registration_open"])
        self.assertTrue(data["collect_phone"])
        self.assertTrue(data["verify_phone"])
        self.assertEqual(data["total_registrations"], 0)
        ticket_names = [t["name"] for t in data["tickets"]]
        self.assertIn("GA", ticket_names)
        self.assertIn("VIP", ticket_names)
