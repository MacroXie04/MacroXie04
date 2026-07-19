"""Excel import implementation for members."""

from __future__ import annotations

from collections.abc import Callable

from django.contrib.auth.hashers import make_password
from django.db import transaction
from django.utils import timezone

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - guarded at runtime
    load_workbook = None

from apps.authn.models import ContactEmail, ContactPhone, Member
from apps.authn.services.contacts.contact_phones import infer_region_from_e164, normalize_to_national

from .operations import bulk_update_members
from .parsing import generate_random_password, normalize_header, parse_row
from .types import ImportResult

BATCH_SIZE = 500


def import_members_from_excel(
    file,
    default_password: str | None = None,
    update_existing: bool = False,
    update_member_allowed: Callable[[Member], bool] | None = None,
) -> ImportResult:
    """Import members from an Excel file using bulk inserts when possible."""
    if load_workbook is None:
        return ImportResult(success=False, errors=["openpyxl library not installed. Please run: pip install openpyxl"])

    result = ImportResult(success=True)
    try:
        workbook = load_workbook(filename=file, read_only=True, data_only=True)
        worksheet = workbook.active
        raw_rows = list(worksheet.iter_rows(values_only=True))
        workbook.close()
        if not raw_rows:
            return ImportResult(success=False, errors=["Excel file is empty"])

        headers = [normalize_header(header) for header in raw_rows[0]]
        if "primary_email" not in headers:
            return ImportResult(success=False, errors=["Excel file must contain a 'Primary Email' column"])

        parsed_rows = []
        seen_in_file: set[str] = set()
        for row_num, row in enumerate(raw_rows[1:], start=2):
            if not any(row):
                continue
            parsed = parse_row(row_num, dict(zip(headers, row, strict=False)))
            if not parsed["primary_email"]:
                result.skipped_count += 1
                result.errors.append(f"Row {row_num}: Missing primary email")
                continue

            email_key = parsed["primary_email"].lower()
            if email_key in seen_in_file:
                result.skipped_count += 1
                result.errors.append(f"Row {row_num}: Duplicate email in file ({parsed['primary_email']})")
                continue
            seen_in_file.add(email_key)
            parsed_rows.append(parsed)

        if not parsed_rows:
            return result

        existing_emails = {email.lower() for email in ContactEmail.objects.values_list("email_address", flat=True)}
        existing_phones = set(ContactPhone.objects.values_list("phone_number", flat=True))
        hashed_pw = make_password(default_password or generate_random_password())
        members_to_create: list[Member] = []
        emails_to_create: list[ContactEmail] = []
        phones_to_create: list[ContactPhone] = []
        rows_to_update: list[dict] = []
        claimed_contact_emails: set[str] = set(existing_emails)
        claimed_phones: set[str] = set(existing_phones)
        now = timezone.now()

        for parsed in parsed_rows:
            email_key = parsed["primary_email"].lower()
            if email_key in existing_emails:
                if update_existing:
                    rows_to_update.append(parsed)
                else:
                    result.skipped_count += 1
                    result.errors.append(
                        f"Row {parsed['row']}: Member with email {parsed['primary_email']} already exists"
                    )
                continue

            if not parsed["first_name"]:
                result.skipped_count += 1
                result.errors.append(f"Row {parsed['row']}: Missing first name")
                continue

            if not parsed["last_name"]:
                result.skipped_count += 1
                result.errors.append(f"Row {parsed['row']}: Missing last name")
                continue

            member = Member(
                password=hashed_pw,
                first_name=parsed["first_name"],
                last_name=parsed["last_name"],
                middle_name=parsed["middle_name"] or None,
                title=parsed["title"] or None,
                organization=parsed["organization"],
                is_active=parsed["is_active"] if parsed["is_active"] is not None else True,
                date_joined=parsed["date_joined"] or now,
            )
            members_to_create.append(member)
            existing_emails.add(email_key)
            _append_contact_records(
                member, parsed, emails_to_create, phones_to_create, claimed_contact_emails, claimed_phones
            )
            result.created_count += 1

        with transaction.atomic():
            # bulk_create bypasses post_save signals, so the Google-Sheet sync
            # signal handlers in authn/signals.py will not fire here. Admins
            # should run the "Sync ALL members to Google Sheet" admin action
            # after import if auto-sync is enabled.
            if members_to_create:
                Member.objects.bulk_create(members_to_create, batch_size=BATCH_SIZE)
            if emails_to_create:
                ContactEmail.objects.bulk_create(emails_to_create, batch_size=BATCH_SIZE)
            if phones_to_create:
                ContactPhone.objects.bulk_create(phones_to_create, batch_size=BATCH_SIZE)
    except Exception as exc:  # noqa: BLE001
        result.success = False
        result.created_count = 0
        result.errors.append(f"Error creating members: {exc}")
        return result

    if rows_to_update:
        # Rebuild claimed sets from actual DB state after successful create
        claimed_contact_emails = {e.lower() for e in ContactEmail.objects.values_list("email_address", flat=True)}
        claimed_phones = set(ContactPhone.objects.values_list("phone_number", flat=True))
        bulk_update_members(rows_to_update, result, claimed_contact_emails, claimed_phones, update_member_allowed)

    return result


def _append_contact_records(member, parsed, emails_to_create, phones_to_create, claimed_contact_emails, claimed_phones):
    email_key = parsed["primary_email"].lower()
    if email_key not in claimed_contact_emails:
        emails_to_create.append(
            ContactEmail(
                member=member,
                email_address=parsed["primary_email"],
                email_type="primary",
                verified=parsed["primary_verified"],
                subscribe=parsed["primary_subscribed"],
            )
        )
        claimed_contact_emails.add(email_key)

    if parsed["secondary_email"]:
        secondary_key = parsed["secondary_email"].lower()
        if secondary_key not in claimed_contact_emails:
            emails_to_create.append(
                ContactEmail(
                    member=member,
                    email_address=parsed["secondary_email"],
                    email_type="secondary",
                    verified=parsed["secondary_verified"],
                    subscribe=parsed["secondary_subscribed"],
                )
            )
            claimed_contact_emails.add(secondary_key)

    if parsed["phone_number"]:
        region = infer_region_from_e164(parsed["phone_number"])
        national = normalize_to_national(parsed["phone_number"], region)
        if national not in claimed_phones:
            phones_to_create.append(
                ContactPhone(
                    member=member,
                    phone_number=national,
                    region=region,
                    subscribe=parsed["phone_subscribed"],
                    verified=parsed["phone_verified"],
                )
            )
            claimed_phones.add(national)
